import time

from deepeval.red_teaming import RedTeamer
from deepeval.models.gpt_model_schematic import SchematicGPTModel
from openpyxl import Workbook
from datetime import datetime
from deepeval.test_case import LLMTestCase
from deepeval.metrics import BaseMetric
from dda.helper import (
    get_titles,
    translate,
    get_attack_prompt,
    get_user_prompt,
    generate_attack_prompts,
    merge_files,
    get_target_model,
)
from dda.red_team_config import params, custom_params


def save_results(df, path, params):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("result")
    write_ws = write_wb.active

    df_list = df.to_dict(orient="records")

    write_ws.append(get_titles())

    # vulnerability_str = ", ".join(str(e) for e in params.get("vulnerabilities"))
    attack_str = ", ".join(
        str(e) for e in list(params.get("attack_enhancements").items())
    )

    for data in df_list:
        vulnerability = str(data.get("Vulnerability", ""))

        input = translate(data.get("Input", ""))
        target = translate(data.get("Target Output", ""))
        reason = translate(data.get("Reason", ""))
        score = data.get("Score", "")

        line = [
            params.get("category"),
            params.get("subcategory"),
            params.get("keywords"),
            params.get("target_purpose"),
            params.get("target_system_prompt"),
            params.get("target_model"),
            params.get("evaluation_model"),
            params.get("synthesizer_model"),
            attack_str,
            vulnerability,
            input,
            target,
            score,
            reason,
        ]
        write_ws.append(line)
    write_wb.save(f"{path}.xlsx")


def main(params):
    """deepeval redteamer를 활용하여 LLM 모델을 공격한 결과데이터를 엑셀파일로 저장하는 함수"""
    start = time.time()

    now = datetime.now().strftime("%y%m%d%H%M")
    file_name = f"breakdown_{params['target_model']}_{params['keywords']}_{now}"
    path = f"./dda/data/{file_name}"

    target_model = get_target_model(params["target_model"])
    if target_model is None:
        raise ValueError(f"Unsupported target model: {params['target_model']}")

    red_teamer = RedTeamer(
        target_purpose=params["target_purpose"],
        target_system_prompt=params["target_system_prompt"],
        evaluation_model=SchematicGPTModel(params["evaluation_model"]),
        synthesizer_model=SchematicGPTModel(params["synthesizer_model"]),
    )
    results = red_teamer.scan(
        target_model=target_model,
        attacks_per_vulnerability=params["attacks_per_vulnerability"],
        attack_enhancements=params["attack_enhancements"],
        vulnerabilities=params["vulnerabilities"],
    )
    df = red_teamer.vulnerability_scores_breakdown
    print(df)
    save_results(df, path, params)

    end = time.time()
    print(f"{end - start:.2f} sec")


def custom_red_teaming(path, params):
    # Create excel file
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("result")
    write_ws = write_wb.active
    write_ws.append(get_titles())

    # Set up red_teamer for scoring
    red_teamer = RedTeamer(
        target_purpose=params["target_purpose"],
        target_system_prompt=params["target_system_prompt"],
        evaluation_model=SchematicGPTModel("gpt-4o"),
        synthesizer_model=SchematicGPTModel("gpt-4o"),
    )
    v_metrics_map = red_teamer.get_red_teaming_metrics_map()

    for idx, v in enumerate(params["vulnerabilities"]):
        metric: BaseMetric = v_metrics_map.get(v)()

        # Generate attack prompts
        system_prompt = get_attack_prompt(
            params,
            "./dda/prompts/red_team_system.txt",
            params["attacks_per_vulnerability"],
        )
        prompt = get_user_prompt(
            "./dda/prompts/red_team_user.txt",
            {
                "category": params["category"],
                "subcategory": params["subcategory"],
                "keywords": params["keywords"],
                "vulnerability": params["vulnerabilities_desc"][idx],
                "target_purpose": params["target_purpose"],
                "target_system_prompt": params["target_system_prompt"],
            },
        )
        results = generate_attack_prompts(
            system_prompt,
            prompt,
            params["target_model"],
        )

        for result in results:
            question = result.get("question")
            answer = result.get("answer")

            test_case = LLMTestCase(
                input=question,
                actual_output=answer,
            )
            metric.measure(test_case)

            line = [
                params.get("category"),
                params.get("subcategory"),
                params.get("keywords"),
                params.get("target_purpose"),
                params.get("target_system_prompt"),
                params.get("target_model"),
                params.get("evaluation_model"),
                params.get("synthesizer_model"),
                "PROMPT_INJECTION",
                str(v),
                question,
                answer,
                metric.score,
                translate(metric.reason),
            ]
            print(f"Complete! {line}")

            write_ws.append(line)
    write_wb.save(f"{path}.xlsx")


def custom_main(params):
    now = datetime.now().strftime("%y%m%d%H%M")
    file_name = f"breakdown_{params['target_model']}_{params['category']}_{params['subcategory']}_{params['keywords']}_{now}"
    path = f"./dda/data/{file_name}"
    custom_red_teaming(path, params)


if __name__ == "__main__":
    # 1) DeepEval Redteamer scanning
    main(params)

    # 2) Custom
    # custom_main(custom_params)

    # If you want to merge excel data, using merge_files function.
    # root_path = f"./dda/data"
    # merge_files(root_path, get_titles())
