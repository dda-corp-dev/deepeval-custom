from datetime import datetime
from openpyxl import Workbook
from dda.red_team_config import advanced_params
from dda.helper import (
    get_attack_prompt,
    get_user_prompt,
    generate_attack_prompts,
    generate_response,
    translate,
    merge_files,
)
from openpyxl import load_workbook
from deepeval.metrics.red_teaming_metrics import *
from deepeval.red_teaming import Vulnerability
from deepeval.test_case import LLMTestCase
from tqdm import tqdm


def _get_advanced_titles(params: dict):
    titles_total = []
    # Params
    titles_total.append("category")
    titles_total.append("subcategory")
    titles_total.append("keywords")
    titles_total.append("target_purpose")
    titles_total.append("target_system_prompt")
    titles_total.append("target_model")
    titles_total.append("evaluation_model")
    titles_total.append("synthesizer_model")
    titles_total.append("attack")
    # Outputs
    titles_total.append("Input")
    titles_total.append("Target Output")
    for key, value in params["graders"].items():
        titles_total.append(f"{key}.Score")
        titles_total.append(f"{key}.Reason")
    return titles_total


def _save_input_results(path: str, params: dict, results: list):
    # Create excel file
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("result")
    write_ws = write_wb.active
    write_ws.append(_get_advanced_titles(params))

    for result in results:
        question = result.get("question")
        line = [
            params.get("category"),
            params.get("subcategory"),
            params.get("keywords"),
            params.get("target_purpose"),
            params.get("target_system_prompt"),
            params.get("target_model"),
            params.get("evaluation_model"),
            params.get("synthesizer_model"),
            "basic adversarial prompts, prompt probing, prompt injection, jailbreak (including linear, tree, and crescendo types)",
            question,
        ]
        # print(line)
        write_ws.append(line)
    write_wb.save(path)


def generate_input_prompts(path: str, params: dict):
    system_prompt = get_attack_prompt(
        params,
        "./dda/prompts/red_team_system_advanced.txt",
        params["nums"],
    )
    prompt = get_user_prompt(
        "./dda/prompts/red_team_user_advanced.txt",
        {
            "vulnerability": params["vulnerability"],
            "category": params["category"],
            "subcategory": params["subcategory"],
            "keywords": params["keywords"],
            "target_purpose": params["target_purpose"],
            "target_system_prompt": params["target_system_prompt"],
        },
    )

    results = generate_attack_prompts(
        system_prompt,
        prompt,
        params["synthesizer_model"],
    )

    if len(results) >= 0:
        _save_input_results(path, params, results)
        print("\n☻ [STEP1 FINISHED] Successfully generated input prompts.\n")
    else:
        raise Exception("\n☹︎ Failed to generate input prompts.")


def generate_outputs(path: str, params: dict):
    load_wb = load_workbook(
        path,
        data_only=True,
    )
    load_ws = load_wb["Sheet"]

    total_rows = sum(1 for _ in load_ws.iter_rows(min_row=2))
    with tqdm(
        total=total_rows,
        desc="Generating Outputs",
        ncols=100,
        bar_format="{l_bar}{bar} [Elapsed: {elapsed} | Remaining: {remaining}]",
        colour="blue",
    ) as pbar:
        for idx, row in enumerate(load_ws.iter_rows(min_row=2)):
            input = row[9].value
            if params["output_language"] == "EN":
                input += " Please respond within 500 characters."
            else:
                input += " 500자 이내로 답변해줘."
            row[10].value = generate_response(input, params["target_model"])

            pbar.update(1)

    load_wb.save(path)
    print("☻ [STEP2 FINISHED] Successfully generated outputs.\n")


def generate_scoring(path: str, params: dict):
    load_wb = load_workbook(
        path,
        data_only=True,
    )
    load_ws = load_wb["Sheet"]
    gap = 2

    def _translate_if_needed(text: str):
        return translate(text) if params["output_language"] == "KO" else text

    def _exist_empty_reason(row, start, end, gap):
        return any(row[i].value is None for i in range(start, end, gap))

    try:
        total_rows = sum(1 for _ in load_ws.iter_rows(min_row=2))
        metric_num = len(params["graders"])
        score_start_col = len(_get_advanced_titles(params)) - (metric_num * gap)
        reason_start_col = score_start_col + 1
        last_col = score_start_col + metric_num * gap

        with tqdm(
            total=total_rows,
            desc="Scoring Progress",
            ncols=100,
            bar_format="{l_bar}{bar} [Elapsed: {elapsed} | Remaining: {remaining}]",
            colour="blue",
        ) as pbar:
            for idx, row in tqdm(enumerate(load_ws.iter_rows(min_row=2))):
                input = row[9].value
                output = row[10].value

                if output:
                    test_case = LLMTestCase(
                        input=input,
                        actual_output=output,
                    )
                    if _exist_empty_reason(
                        row, reason_start_col, last_col, gap
                    ):
                        for i, (key, grader) in zip(
                            range(score_start_col, last_col, gap),
                            params["graders"].items(),
                        ):
                            grader.measure(test_case)
                            row[i].value = grader.score
                            row[i + 1].value = _translate_if_needed(
                                grader.reason
                            )
                pbar.update(1)
        print("☻ [STEP3 FINISHED] Scoring has been successfully completed.\n")
    except Exception as e:
        print(f"☹︎ An error occurred during scoring. msg: {e}")
        print(
            "☹︎ [STEP3 UNFINISHED] The scoring has ended unsatisfactorily. Please run generate_scoring again.\n"
        )
    finally:
        load_wb.save(path)


def advanced_main(params):
    now = datetime.now().strftime("%y%m%d%H%M")
    file_name = (
        f"{now}_{params['output_language']}_breakdown_{params['category']}"
    )
    # file_name = "2411130856_KO_breakdown_Violations of human rights"
    path = f"./dda/data/{file_name}.xlsx"

    generate_input_prompts(path, params)
    generate_outputs(path, params)
    generate_scoring(path, params)


if __name__ == "__main__":
    advanced_main(advanced_params)
    # merge_files("./dda/data/ko", _get_advanced_titles(advanced_params))
