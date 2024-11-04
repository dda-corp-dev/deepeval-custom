import os
import glob
import re
import deepl
import requests
import json
import transformers
import openai
import logging


from typing import Optional, Tuple
from deepeval.models import DeepEvalBaseLLM
from anthropic import Anthropic
from transformers import AutoModelForCausalLM, AutoTokenizer
from openpyxl import load_workbook, Workbook
from deepeval.key_handler import KeyValues, KEY_FILE_HANDLER
from anthropic import Anthropic
from langchain_openai import ChatOpenAI, AzureChatOpenAI
from langchain_community.callbacks import get_openai_callback
from tenacity import retry, retry_if_exception_type, wait_exponential_jitter


valid_gpt_models = [
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-4-turbo",
    "gpt-4-turbo-preview",
    "gpt-4-0125-preview",
    "gpt-4-1106-preview",
    "gpt-4",
    "gpt-4-32k",
    "gpt-4-0613",
    "gpt-4-32k-0613",
    "gpt-3.5-turbo-1106",
    "gpt-3.5-turbo",
    "gpt-3.5-turbo-16k",
    "gpt-3.5-turbo-0125",
]

default_gpt_model = "gpt-4o"


def log_retry_error(retry_state):
    logging.error(
        f"OpenAI rate limit exceeded. Retrying: {retry_state.attempt_number} time(s)..."
    )


class TargetGPTModel(DeepEvalBaseLLM):
    def __init__(
        self,
        model: Optional[str] = None,
        _openai_api_key: Optional[str] = None,
        *args,
        **kwargs,
    ):
        model_name = None
        if isinstance(model, str):
            model_name = model
            if model_name not in valid_gpt_models:
                raise ValueError(
                    f"Invalid model. Available GPT models: {', '.join(model for model in valid_gpt_models)}"
                )
        elif model is None:
            model_name = default_gpt_model

        self._openai_api_key = _openai_api_key
        # args and kwargs will be passed to the underlying model, in load_model function
        self.args = args
        self.kwargs = kwargs
        super().__init__(model_name)

    def load_model(self):
        if self.should_use_azure_openai():
            openai_api_key = KEY_FILE_HANDLER.fetch_data(
                KeyValues.AZURE_OPENAI_API_KEY
            )

            openai_api_version = KEY_FILE_HANDLER.fetch_data(
                KeyValues.OPENAI_API_VERSION
            )
            azure_deployment = KEY_FILE_HANDLER.fetch_data(
                KeyValues.AZURE_DEPLOYMENT_NAME
            )
            azure_endpoint = KEY_FILE_HANDLER.fetch_data(
                KeyValues.AZURE_OPENAI_ENDPOINT
            )

            model_version = KEY_FILE_HANDLER.fetch_data(
                KeyValues.AZURE_MODEL_VERSION
            )

            if model_version is None:
                model_version = ""

            return AzureChatOpenAI(
                openai_api_version=openai_api_version,
                azure_deployment=azure_deployment,
                azure_endpoint=azure_endpoint,
                openai_api_key=openai_api_key,
                model_version=model_version,
                *self.args,
                **self.kwargs,
            )

        return ChatOpenAI(
            model_name=self.model_name,
            openai_api_key=self._openai_api_key,
            *self.args,
            **self.kwargs,
        )

    @retry(
        wait=wait_exponential_jitter(initial=1, exp_base=2, jitter=2, max=10),
        retry=retry_if_exception_type(openai.RateLimitError),
        after=log_retry_error,
    )
    def generate(self, prompt: str) -> Tuple[str, float]:
        chat_model = self.load_model()
        with get_openai_callback() as cb:
            res = chat_model.invoke(prompt)
            return res.content

    @retry(
        wait=wait_exponential_jitter(initial=1, exp_base=2, jitter=2, max=10),
        retry=retry_if_exception_type(openai.RateLimitError),
        after=log_retry_error,
    )
    async def a_generate(self, prompt: str) -> Tuple[str, float]:
        chat_model = self.load_model()
        with get_openai_callback() as cb:
            res = await chat_model.ainvoke(prompt)
            return res.content

    def should_use_azure_openai(self):
        value = KEY_FILE_HANDLER.fetch_data(KeyValues.USE_AZURE_OPENAI)
        return value.lower() == "yes" if value is not None else False

    def get_model_name(self):
        if self.should_use_azure_openai():
            return "azure openai"
        elif self.model_name:
            return self.model_name


class CustomClaudeOpus(DeepEvalBaseLLM):
    def __init__(self):
        self.model = Anthropic(
            api_key=KEY_FILE_HANDLER.fetch_data(KeyValues.ANTHROPIC_API_KEY)
        )

    def load_model(self):
        return self.model

    def generate(self, prompt: str):
        client = self.load_model()
        resp = client.messages.create(
            model="claude-3-opus-20240229",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
        )
        return resp.content[0].text

    async def a_generate(self, prompt: str):
        return self.generate(prompt)

    def get_model_name(self):
        return "Clause-3 Opus"


class CustomQwen2_3B(DeepEvalBaseLLM):
    def __init__(self):
        model_name = "Qwen/Qwen2.5-3B-Instruct"
        model = AutoModelForCausalLM.from_pretrained(
            model_name, torch_dtype="auto", device_map="cpu"
        )
        tokenizer = AutoTokenizer.from_pretrained(model_name)

        self.model = model
        self.tokenizer = tokenizer

    def load_model(self):
        return self.model

    def generate(self, prompt: str) -> str:
        model = self.load_model()
        messages = [
            {
                "role": "system",
                "content": "You are an AI Korean assistant designed to fulfill the user's spoken requests.",
            },
            {"role": "user", "content": prompt},
        ]
        text = self.tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )
        pipeline = transformers.pipeline(
            "text-generation",
            model=model,
            tokenizer=self.tokenizer,
            use_cache=True,
            device_map="cpu",
            max_length=2048,
            do_sample=True,
            top_k=5,
            num_return_sequences=1,
        )
        result = pipeline(text)
        return result[0]["generated_text"].split("<|im_start|>assistant\n")[-1]

    async def a_generate(self, prompt: str) -> str:
        return self.generate(prompt)

    def get_model_name(self):
        return "QWen-2.5 3B"


class CustomLlama3_1B(DeepEvalBaseLLM):
    def __init__(self):
        model_name = "meta-llama/Llama-3.2-1B-Instruct"
        model = AutoModelForCausalLM.from_pretrained(
            model_name,
            torch_dtype="auto",
            device_map="cpu",
        )
        tokenizer = AutoTokenizer.from_pretrained(model_name)

        self.model = model
        self.tokenizer = tokenizer

    def load_model(self):
        return self.model

    def generate(self, prompt: str) -> str:
        model = self.load_model()
        messages = [
            {
                "role": "system",
                "content": "You are an AI Korean assistant designed to fulfill the user's spoken requests.",
            },
            {"role": "user", "content": prompt},
        ]
        pipe = transformers.pipeline(
            "text-generation",
            model=model,
            tokenizer=self.tokenizer,
            use_cache=True,
            device_map="auto",
            max_length=1000,
            do_sample=True,
            top_k=5,
            num_return_sequences=1,
        )
        result = pipe(
            messages,
        )
        content = result[0]["generated_text"][-1].get("content", "")
        return content

    async def a_generate(self, prompt: str) -> str:
        return self.generate(prompt)

    def get_model_name(self):
        return "Llama-3.1 1B"


def get_target_model(name):
    if name == "gpt-3.5-turbo-0125":
        return TargetGPTModel(name)
    elif name == "claude-3-opus-20240229":
        return CustomClaudeOpus()
    elif name == "qwen2-3b":
        return CustomQwen2_3B()
    elif name == "llama3-1b":
        return CustomLlama3_1B()
    return None


def merge_files(root_path):
    """root_path 안에 있는 모든 excel파일을 합쳐서 하나의 최종 파일로 생성하는 함수"""
    files = glob.glob(os.path.join(root_path, "**", "*.xlsx"), recursive=True)

    final_wb = Workbook()
    final_ws = final_wb.active

    title = get_titles()
    final_ws.append(title)

    for file in files:
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            final_ws.append(row)

    final_wb.save(f"{root_path}/merged.xlsx")


def get_titles():
    titles_total = []
    # Params
    titles_total.append("category")
    titles_total.append("subcategory")
    titles_total.append("keywords")
    titles_total.append("target_purpose")
    titles_total.append("target_system_prompt")
    titles_total.append("target_model")
    titles_total.append("evaluation_model")
    titles_total.append("synthesizer_model")
    titles_total.append("attack")
    # Outputs
    titles_total.append("Vulnerability")
    titles_total.append("Input")
    titles_total.append("Target Output")
    titles_total.append("Score")
    titles_total.append("Reason")
    return titles_total


def _contain_english_string(input):
    # 정규식 패턴: 영어 알파벳이 하나 이상 포함되어 있는지 검사
    eng_regex = re.compile(r"[a-zA-Z]")
    return bool(eng_regex.search(input))


def _translate_deepl(sentence: str, target: str):
    auth_key = KEY_FILE_HANDLER.fetch_data(KeyValues.DEEPL_AUTH_KEY)
    translator = deepl.Translator(auth_key)
    result = translator.translate_text(sentence, target_lang=target)
    return result.text


def translate(input):
    is_english = _contain_english_string(input)
    if is_english:
        return _translate_deepl(input, "KO")
    else:
        return input


def generate_attack_prompts(
    system_prompt: str,
    user_prompt: str,
    model_name: str,
):
    results = []
    try:
        if model_name.find("claude") != -1:
            messages = [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": user_prompt,
                        }
                    ],
                }
            ]
            inputs = get_claude3_answer(
                model_name,
                system_prompt,
                messages,
                0.4,
                8000,
            )
            results = json.loads(inputs).get("result")
        elif model_name.find("gpt") != -1:
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ]
            results = get_azure_gpt_answer(
                messages,
                0.4,
                4000,
            )
            results = json.loads(results).get("result")
    except Exception as e:
        print(f"[API ERROR] {e}")
    return results


def get_claude3_answer(
    model: str,
    system: str,
    messages: list,
    temperature: int,
    max_tokens: int,
) -> str:
    try:
        client = Anthropic(
            api_key=KEY_FILE_HANDLER.fetch_data(KeyValues.ANTHROPIC_API_KEY)
        )
        message = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            temperature=temperature,
            system=system,
            messages=messages,
        )
    except Exception as e:
        print(f"get_claude3_answer error: {e}")
    else:
        return message.content[0].text


def get_azure_gpt_answer(
    messages: list, temperature: float, max_tokens: int
) -> str:
    try:
        AZURE_KEY = KEY_FILE_HANDLER.fetch_data(KeyValues.AZURE_OPENAI_API_KEY)
        AZURE_API_URL = KEY_FILE_HANDLER.fetch_data(
            KeyValues.AZURE_OPENAI_ENDPOINT
        )
        headers = {
            "Content-Type": "application/json",
            "api-key": AZURE_KEY,
        }
        payload = {
            "messages": messages,
            "temperature": temperature,
            "top_p": 0.95,
            "max_tokens": max_tokens,
            "response_format": {"type": "json_object"},
        }
        response = requests.post(
            AZURE_API_URL,
            headers=headers,
            json=payload,
        )
        response.raise_for_status()
        result = response.json()
    except requests.RequestException as e:
        print(f"[API REQUEST EXCEPTION] {e}")
    except Exception as e:
        print(f"[API ERROR] {e}")
    else:
        return result["choices"][0]["message"]["content"]


def get_attack_prompt(path: str, num: int):
    with open(path, "r", encoding="utf-8") as f:
        prompt = f.read()
    prompt = prompt.replace("#{num}", str(num))
    return prompt


def get_user_prompt(path: str, params: dict):
    with open(path, "r", encoding="utf-8") as f:
        prompt = f.read()
    for key in params.keys():
        key_str = f"#{{{key}}}"
        prompt = prompt.replace(key_str, params[key])
    return prompt
